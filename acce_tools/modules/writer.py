"""writer.py — Writes ACCERecord lists to .xlsx, .csv, or .json files.

Single responsibility: serialize a list of ACCERecord objects to disk.
Output format is detected from the file extension of output_path.
All formats produce the same unified column set in the same order.
"""

import csv
import json
import os
from typing import Callable

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .schema import ACCERecord


# ── Unified column definition ─────────────────────────────────────
# Order follows ACCE Spreadsheet Import/Export conventions.
# Each entry: (column_header, getter_function)

COLUMNS: list[tuple[str, Callable[[ACCERecord], object]]] = [
    ("USER_TAG",              lambda r: r.user_tag),
    ("DESCRIPTION",           lambda r: r.description_en or r.description_ru),
    ("DESCRIPTION_RU",        lambda r: r.description_ru),
    ("PARENT_AREA",           lambda r: r.parent_area),
    ("ACTION",                lambda r: r.action),
    ("FLOW_RATE",             lambda r: r.flow_rate),
    ("FLOW_RATE_UNIT",        lambda r: r.flow_rate_unit),
    ("PRESSURE",              lambda r: r.pressure),
    ("PRESSURE_UNIT",         lambda r: r.pressure_unit),
    ("DESIGN_TEMPERATURE",    lambda r: r.design_temperature),
    ("OPERATING_TEMPERATURE", lambda r: r.operating_temperature),
    ("DIAMETER",              lambda r: r.diameter_m),
    ("LENGTH",                lambda r: r.length_m),
    ("VOLUME",                lambda r: r.volume),
    ("MOTOR_POWER_KW",        lambda r: r.motor_power_kw),
    ("CAPACITY_KW",           lambda r: r.capacity_kw),
    ("HEAT_TRANSFER_AREA",    lambda r: r.heat_transfer_area_m2),
    ("HEAT_DUTY",             lambda r: r.heat_duty_gcalh),
    ("WEIGHT_KG",             lambda r: r.weight_unit_kg),
    ("WEIGHT_TOTAL_KG",       lambda r: r.weight_total_kg),
    ("DN_MM",                 lambda r: r.dn_mm),
    ("LIFT_CAPACITY_T",       lambda r: r.lift_capacity_t),
    ("VFD",                   lambda r: "YES" if r.vfd else ""),
    ("EXPLOSION_PROOF",       lambda r: "YES" if r.explosion_proof else ""),
    ("OPERATION_STATUS",      lambda r: r.operation_status),
    ("QUANTITY",              lambda r: r.quantity),
    ("MATERIAL",              lambda r: r.material),
    ("ACCE_ITEM_SYMBOL",       lambda r: r.acce_item_symbol),
    ("ACCE_ITEM_TYPE",         lambda r: r.acce_item_type),
    ("CLASSIFICATION_SOURCE",  lambda r: r.classification_source),
    ("RAW_TAG",               lambda r: r.raw_tag),
    ("TECH_RAW",              lambda r: r.tech_raw),
    ("SOURCE_FILE",           lambda r: r.source_file),
    ("SOURCE_SHEET",          lambda r: r.source_sheet),
    ("SOURCE_ROW",            lambda r: r.source_row),
    ("IS_VALID",              lambda r: "YES" if r.is_valid else "NO"),
    ("ERRORS",                lambda r: "; ".join(r.errors) if r.errors else ""),
    ("WARNINGS",              lambda r: "; ".join(str(w) for w in r.warnings) if r.warnings else ""),
]


def _coerce(v: object) -> object:
    """
    Normalizes a field value for output:
      None, float NaN  → ""
      bool             → already handled by getters (returns "YES" or "")
      everything else  → unchanged (keeps int/float as numeric for xlsx)
    """
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN check
        return ""
    s = str(v)
    if s in ("None", "nan", "NaN"):
        return ""
    return v


def _row_values(record: ACCERecord) -> list:
    return [_coerce(getter(record)) for _, getter in COLUMNS]


# ── xlsx ─────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_xlsx(records: list[ACCERecord], output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACCE_Import"

    headers = [col for col, _ in COLUMNS]
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.fill  = _HEADER_FILL
        cell.font  = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"

    for record in records:
        ws.append(_row_values(record))

    # Auto-size columns (cap at 60 characters)
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_width = max(
            len(str(cell.value or ""))
            for cell in ws[col_letter]
        )
        ws.column_dimensions[col_letter].width = min(max_width + 2, 60)

    wb.save(output_path)


# ── csv ──────────────────────────────────────────────────────────

def _write_csv(records: list[ACCERecord], output_path: str) -> None:
    headers = [col for col, _ in COLUMNS]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for record in records:
            writer.writerow([str(v) if v != "" else "" for v in _row_values(record)])


# ── json ─────────────────────────────────────────────────────────

def _write_json(records: list[ACCERecord], output_path: str) -> None:
    data = []
    for record in records:
        row = {}
        for (col_name, _), v in zip(COLUMNS, _row_values(record)):
            row[col_name] = v if v != "" else ""
        data.append(row)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


# ── PUBLIC API ────────────────────────────────────────────────────

_WRITERS: dict[str, Callable] = {
    ".xlsx": _write_xlsx,
    ".csv":  _write_csv,
    ".json": _write_json,
}


def write(records: list[ACCERecord], output_path: str) -> None:
    """
    Writes a list of ACCERecord objects to output_path.
    Format is detected from the file extension (.xlsx, .csv, .json).
    """
    ext = os.path.splitext(output_path)[1].lower()
    writer_fn = _WRITERS.get(ext)
    if writer_fn is None:
        raise ValueError(
            f"Unsupported output extension '{ext}'. "
            f"Supported: {', '.join(_WRITERS)}"
        )
    writer_fn(records, output_path)
