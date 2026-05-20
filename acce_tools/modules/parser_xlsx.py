"""parser_xlsx.py — Reads an Excel equipment list and returns raw ACCERecord objects.

Single responsibility: open the file, extract rows, map cells to ACCERecord fields.
Classification (acce_item_symbol / acce_item_type) is NOT performed here.
"""

import re
import openpyxl
from typing import Optional

from .schema import ACCERecord
from .spec_parser import parse_spec


# ─────────────────────────────────────────────────────────────────
# ПАРСИНГ ТЕХ. ХАРАКТЕРИСТИК
# Обрабатывает паттерны: Q=160m3/h  P=0.8MPa  V=6m3  Φ3.8×25m  DN500
# ─────────────────────────────────────────────────────────────────

_UNITS = r'(m3/h|m3|kW|MPa|kPa|Pa|t/h|kg/h|t|m)?'


def _extract_param(text: str, key: str) -> tuple[Optional[float], str]:
    """Extracts Q=, P=, V= etc. For ranges takes the lower value."""
    pattern = (
        rf'(?<![A-Za-z]){re.escape(key)}\s*=\s*'
        rf'([0-9][0-9,\.]*)'
        rf'(?:\s*[-~]\s*[0-9][0-9,\.]*)?'
        rf'\s*' + _UNITS
    )
    m = re.search(pattern, text, re.IGNORECASE)
    if not m:
        return None, ""
    raw = m.group(1).replace(',', '')
    unit = m.group(2) or ""
    try:
        return float(raw), unit
    except ValueError:
        return None, ""


def _extract_dn(text: str) -> Optional[int]:
    m = re.search(r'DN\s*(\d+)', text, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _extract_phi(text: str) -> tuple[Optional[float], Optional[float]]:
    """Φ3.8×25m → (diameter_m=3.8, length_m=25.0)"""
    m = re.search(r'[ΦΦ]?\s*([\d.]+)\s*[×x]\s*([\d.]+)', text)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _extract_lift(text: str) -> Optional[float]:
    """T=5t → 5.0 (crane/hoist rated load in tonnes)"""
    m = re.search(r'T\s*=\s*([\d.]+)\s*t', text)
    return float(m.group(1)) if m else None


def _parse_tech(raw: str) -> dict:
    """Parses a technical specification string into structured fields."""
    if not raw:
        return {}
    result = {}

    q_val, q_unit = _extract_param(raw, 'Q')
    p_val, p_unit = _extract_param(raw, 'P')
    v_val, v_unit = _extract_param(raw, 'V')
    dn             = _extract_dn(raw)
    dia, length    = _extract_phi(raw)
    lift           = _extract_lift(raw)

    if q_val is not None:
        if q_unit and q_unit.lower() == 'kw':
            result['capacity_kw']    = q_val
        else:
            result['flow_rate']      = q_val
            result['flow_rate_unit'] = q_unit

    if p_val is not None:
        result['pressure']      = p_val
        result['pressure_unit'] = p_unit

    if v_val is not None:
        result['volume']      = v_val
        result['volume_unit'] = v_unit

    if dn is not None:
        result['dn_mm'] = dn

    if dia is not None:
        result['diameter_m'] = dia
    if length is not None:
        result['length_m'] = length

    if lift is not None:
        result['lift_capacity_t'] = lift

    return result


# ─────────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────────────────────────────

def _build_merge_map(sheet) -> dict:
    """Maps every cell in a merged range to the top-left cell's value."""
    merge_map = {}
    for rng in sheet.merged_cells.ranges:
        val = sheet.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_map[(r, c)] = val
    return merge_map


def _cell(sheet, row: int, col: int, mm: dict):
    v = sheet.cell(row=row, column=col).value
    return v if v is not None else mm.get((row, col))


def _str(v) -> str:
    if v is None:
        return ""
    return re.sub(r'\s+', ' ', str(v).strip())


def _float(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "n/a", "tbd", "#REF!", "#VALUE!"):
        return None
    try:
        return float(s)
    except ValueError:
        m = re.search(r'-?\d[\d,\.]*', s)
        if m:
            try:
                return float(m.group().replace(',', ''))
            except ValueError:
                pass
    return None


def _int(v) -> Optional[int]:
    f = _float(v)
    return int(f) if f is not None else None


def _split_bilingual(text: str) -> tuple[str, str]:
    """'Насос / Pump' → ('Насос', 'Pump'). No separator → all goes to RU field."""
    if '/' in text:
        parts = text.split('/', 1)
        return parts[0].strip(), parts[1].strip()
    return text.strip(), ""


_FORBIDDEN = re.compile(r'[!@#$%&*()?/{}[\]*=<>,`~^:;|"\' \\+\s～]')


def _make_user_tag(raw_tag: str, row_num: int) -> str:
    """Creates a valid ACCE user_tag: max 20 chars, no spaces or special characters."""
    safe = _FORBIDDEN.sub('-', raw_tag.strip())
    safe = re.sub(r'-{2,}', '-', safe).strip('-')
    safe = safe[:20]
    return safe if safe else f"SEQ-{row_num:04d}"


# ─────────────────────────────────────────────────────────────────
# SHEET AND COLUMN DETECTION
# ─────────────────────────────────────────────────────────────────

_TAG_SIGNALS = ('таговый', 'тэговый', 'тег', 'tag')

_COL_PATTERNS: dict[str, list[str]] = {
    'area':   ['титул', 'зона', 'area', 'установка'],
    'tag':    ['таговый', 'тэговый', 'тег', 'tag'],
    'desc':   ['наименован', 'описание', 'description', 'name'],
    'tech':   ['характеристик', 'тех.', 'тех ', 'spec', 'technical'],
    'qty':    ['кол.', 'количеств', 'qty', 'quantity'],
    'wu':     ['вес единицы', 'масса единицы', 'unit weight', 'unit wt'],
    'wt':     ['масса, кг', 'масса,кг', 'total kg', 'всего кг'],
    'wtt':    ['тонн', 'total t'],
    'motor':  ['мощност', 'power (kw)', 'мощность электро'],
    'vfd':    ['частот', 'frequency convers'],
    'expl':   ['взрыв', 'explosion'],
    'status': ['состояние', 'operation status'],
}

# Fallbacks for required columns only; optional ones default to -1 (absent).
_COL_REQUIRED_DEFAULTS: dict[str, int] = {
    'tag': 3, 'desc': 4, 'tech': 6, 'qty': 7,
}


def _scan_cell(v) -> str:
    return str(v).lower().strip() if v is not None else ""


def _find_sheet(wb):
    """
    Returns (worksheet, header_row_index).
    Scans all sheets for the first row that contains a tag-column keyword.
    Falls back to the first sheet, row 1 if nothing matches.
    """
    for name in wb.sheetnames:
        ws = wb[name]
        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_text = " ".join(
                _scan_cell(ws.cell(row=row_idx, column=c).value)
                for c in range(1, min(ws.max_column + 1, 30))
            )
            if any(sig in row_text for sig in _TAG_SIGNALS):
                return ws, row_idx
    return wb[wb.sheetnames[0]], 1


def _detect_cols(ws, header_row: int) -> dict[str, int]:
    """
    Maps column roles to 1-based column indices by scanning the header row.
    Required roles fall back to hardcoded defaults; optional roles return -1.
    """
    n = min(ws.max_column, 30)
    headers = [
        _scan_cell(ws.cell(row=header_row, column=c).value)
        for c in range(1, n + 1)
    ]
    result: dict[str, int] = {}
    for role, patterns in _COL_PATTERNS.items():
        for ci, h in enumerate(headers, start=1):
            if any(p in h for p in patterns):
                result[role] = ci
                break
        else:
            result[role] = _COL_REQUIRED_DEFAULTS.get(role, -1)
    return result


def _find_data_start(ws, header_row: int) -> int:
    """
    Returns the first row that contains actual data after the header.
    Skips blank rows and sub-header rows (where col 1 is non-numeric).
    Falls back to header_row + 2.
    """
    for row_idx in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
        v = ws.cell(row=row_idx, column=1).value
        if v is None:
            continue
        try:
            int(float(str(v).strip()))
            return row_idx
        except (ValueError, TypeError):
            continue
    return header_row + 2


# ─────────────────────────────────────────────────────────────────
# PUBLIC API
# ─────────────────────────────────────────────────────────────────

def parse(file_path: str) -> list[ACCERecord]:
    """Reads an Excel equipment list and returns a list of ACCERecord objects."""
    import os
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws, header_row = _find_sheet(wb)
    cols = _detect_cols(ws, header_row)
    data_start = _find_data_start(ws, header_row)
    mm = _build_merge_map(ws)
    records: list[ACCERecord] = []

    for row_idx in range(data_start, ws.max_row + 1):
        vals = [_cell(ws, row_idx, c, mm) for c in range(1, 12)]
        if all(v is None for v in vals):
            continue

        area      = _str(_cell(ws, row_idx, cols['area'], mm)) if cols['area'] > 0 else ""
        tag       = _str(_cell(ws, row_idx, cols['tag'],  mm))
        desc_raw  = _str(_cell(ws, row_idx, cols['desc'], mm))
        tech_raw  = _str(_cell(ws, row_idx, cols['tech'], mm))
        qty       = _int(_cell(ws, row_idx, cols['qty'],  mm))
        w_unit    = _float(_cell(ws, row_idx, cols['wu'],  mm)) if cols['wu']  > 0 else None
        w_total   = _float(_cell(ws, row_idx, cols['wt'],  mm)) if cols['wt']  > 0 else None
        w_total_t = _float(_cell(ws, row_idx, cols['wtt'], mm)) if cols['wtt'] > 0 else None

        seq = _int(_cell(ws, row_idx, 1, mm))
        if not tag and not seq:
            continue
        if not tag and not desc_raw:
            continue

        row_warns = []
        if not tag:
            row_warns.append("Missing equipment tag")
        if not desc_raw:
            row_warns.append("Missing equipment name")

        desc_ru, desc_en = _split_bilingual(desc_raw)
        tech = _parse_tech(tech_raw)

        record = ACCERecord(
            source_file  = file_path,
            source_sheet = ws.title,
            source_row   = row_idx,

            raw_tag         = tag,
            user_tag        = _make_user_tag(tag, row_idx),
            description_ru  = desc_ru,
            description_en  = desc_en,
            parent_area     = area,
            quantity        = qty or 1,

            weight_unit_kg  = w_unit,
            weight_total_kg = w_total,
            weight_total_t  = w_total_t,

            tech_raw        = tech_raw,
            flow_rate       = tech.get('flow_rate'),
            flow_rate_unit  = tech.get('flow_rate_unit', ''),
            pressure        = tech.get('pressure'),
            pressure_unit   = tech.get('pressure_unit', ''),
            volume          = tech.get('volume'),
            volume_unit     = tech.get('volume_unit', ''),
            capacity_kw     = tech.get('capacity_kw'),
            diameter_m      = tech.get('diameter_m'),
            length_m        = tech.get('length_m'),
            dn_mm           = tech.get('dn_mm'),
            lift_capacity_t = tech.get('lift_capacity_t'),

            action   = "NEW",
            is_valid = len(row_warns) == 0,
            warnings = row_warns,
        )

        # Apply spec_parser results to any fields not yet populated
        for f, v in parse_spec(tech_raw).items():
            if hasattr(record, f) and getattr(record, f) is None:
                setattr(record, f, v)

        # Motor power kW
        motor_col = cols.get('motor', -1)
        if motor_col > 0:
            pwr_raw = str(_cell(ws, row_idx, motor_col, mm) or '')
            m = re.search(r'([\d.]+)', pwr_raw)
            if m and record.motor_power_kw is None:
                record.motor_power_kw = float(m.group(1))

        # VFD
        vfd_col = cols.get('vfd', -1)
        if vfd_col > 0:
            val = str(_cell(ws, row_idx, vfd_col, mm) or '').upper()
            record.vfd = 'ДА' in val or 'YES' in val

        # Explosion proof
        expl_col = cols.get('expl', -1)
        if expl_col > 0:
            val = str(_cell(ws, row_idx, expl_col, mm) or '').upper()
            record.explosion_proof = 'ДА' in val or 'YES' in val

        # Operation status
        status_col = cols.get('status', -1)
        if status_col > 0:
            val = str(_cell(ws, row_idx, status_col, mm) or '').upper()
            if 'РЕЗЕРВ' in val or 'STANDBY' in val or 'SPAR' in val:
                record.operation_status = 'SPAR'
            else:
                record.operation_status = 'DUTY'

        records.append(record)

    return records
