"""Main exporter: converts parser output to ACCE-ready xlsx."""

from __future__ import annotations

import logging
import math
import os
import re
import shutil
import zipfile
from dataclasses import dataclass, field

import pandas as pd
from openpyxl.cell import MergedCell as _MergedCell

from .field_map import MANDATORY_SLOTS, RECORD_COLUMN_MAP, UNIVERSAL_MAP
from .router import route
from .template_reader import SheetSchema, _FIRST_DATA, read_template
from .unit_converter import convert
from .unknown_handler import write_unknown_sheet

logger = logging.getLogger(__name__)

# Default source units for fields that have no separate unit column in
# the parser output (all uppercase to match the normalised CONVERTERS keys).
_SOURCE_UNITS: dict[str, str] = {
    "design_temperature":    "DEG C",
    "operating_temperature": "DEG C",
    "diameter_m":            "M",
    "length_m":              "M",
    "volume":                "M3",
    "motor_power_kw":        "KW",
    "capacity_kw":           "KW",
    "heat_transfer_area_m2": "M2",
    "heat_duty_gcalh":       "GCAL/H",
    "lift_capacity_t":       "T",
}


@dataclass
class ExportSummary:
    total_records: int
    exported: int           # records written to ACCE sheets
    skipped_invalid: int    # IS_VALID = NO (and not overridden by parent_area)
    unknown_symbol: int     # ACCE_ITEM_SYMBOL = UNKNOWN or no route found
    used_default_type: int  # item_type was UNKNOWN, used SYMBOL_DEFAULTS
    by_sheet: dict[str, int] = field(default_factory=dict)


# ── Public API ────────────────────────────────────────────────────────────────

def export(
    input_path: str,
    template_path: str,
    output_path: str,
    default_area: str = "Main Area",
) -> ExportSummary:
    """Convert parser output xlsx to an ACCE-ready xlsx using the master template."""

    # ── Step 1: read template ────────────────────────────────────────────────
    schema = read_template(template_path)
    available_sheets = set(schema.keys())

    # Copy the template file to the output path to avoid style corruption that
    # occurs with copy.deepcopy on openpyxl Workbook objects, then open it.
    shutil.copy2(template_path, output_path)
    import openpyxl
    wb = openpyxl.load_workbook(output_path)

    # Stamp a unique non-empty CpUserTag on the IGNORE service row (row 11) of
    # every equipment sheet.  ACCE includes both the IGNORE row and the LAST ROW
    # sentinel in its UserTag uniqueness check; both default to blank, which
    # triggers "Duplicate user tag number found" on every sheet.  A sheet-
    # specific placeholder makes the two rows distinct.
    _NON_EQ_SHEETS = {'CONTENTS', 'AREAS', 'ScrollData', 'Version',
                      'UNKNOWN', 'SKIPPED'}
    for sheet_name, sc in schema.items():
        if sheet_name in _NON_EQ_SHEETS:
            continue
        if sheet_name not in wb.sheetnames:
            continue
        if 'CpUserTag' in sc.col_index:
            usertag_col = sc.col_index['CpUserTag'] + 1  # 1-based
            wb[sheet_name].cell(row=11, column=usertag_col).value = \
                f'_IGNORE_{sheet_name[:10]}'

    # Prepare every equipment sheet: unmerge decorative merged ranges and clear
    # the original LAST ROW marker so it won't interfere after we place a new one.
    # We do NOT delete or insert rows — writing data directly to the existing
    # template rows keeps GENERAL INSTRUCTIONS and CONTENTS at their original
    # positions, avoiding the openpyxl ghost-cell bug that occurs with large
    # row-shift operations.
    for sheet_name, sc in schema.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Unmerge only ranges that are entirely within the writable data zone
        # [_FIRST_DATA, last_row_index).  The GENERAL INSTRUCTIONS block sits
        # BELOW last_row_index (e.g. rows 1012-1031 while LAST ROW = 1008) and
        # must stay merged — it is the hyperlink "button" back to CONTENTS.
        for merge_range in list(ws.merged_cells.ranges):
            if (merge_range.max_row >= _FIRST_DATA
                    and merge_range.min_row < sc.last_row_index):
                ws.unmerge_cells(str(merge_range))

        # Clear data zone rows 12..last_row_index, columns B onwards.
        # Column A is ACCE-reserved (LAST ROW marker, CONTENTS refs) — never touch it.
        if sc.last_row_index >= _FIRST_DATA:
            for r in range(_FIRST_DATA, sc.last_row_index + 1):
                for c in range(2, ws.max_column + 1):
                    ws.cell(r, c).value = None

    # ── Step 2: read parser output ───────────────────────────────────────────
    df = pd.read_excel(input_path, sheet_name="ACCE_Import", header=0)
    records: list[dict] = df.to_dict(orient="records")
    logger.info("Read %d records from %s", len(records), input_path)

    # ── Step 3: classify records into export / unknown / skipped groups ──────
    export_groups: dict[str, list[dict]] = {}
    unknown_records: list[dict] = []
    skipped_invalid: list[dict] = []
    used_default_type = 0

    for rec in records:
        is_valid = rec.get("IS_VALID")
        valid = is_valid in (True, 'YES', 'yes', 1, '1')

        if not valid:
            skipped_invalid.append(rec)
            continue

        symbol    = str(rec.get("ACCE_ITEM_SYMBOL") or "").strip().upper()
        item_type = str(rec.get("ACCE_ITEM_TYPE")   or "").strip().upper()

        if symbol in ('UNKNOWN', '') or not symbol:
            unknown_records.append(rec)
            continue

        sheet_name, routing_note = route(symbol, item_type, available_sheets)

        if sheet_name is None:
            logger.debug(
                "No route for %s / %s: %s — adding to UNKNOWN",
                symbol, item_type, routing_note,
            )
            unknown_records.append(rec)
            continue

        if "fallback:" in routing_note:
            used_default_type += 1

        export_groups.setdefault(sheet_name, []).append(rec)

    # ── Step 4: write each group into its sheet ──────────────────────────────
    by_sheet: dict[str, int] = {}
    parent_areas: set[str] = set()

    for sheet_name, rec_list in export_groups.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(
                "Sheet %s is not in the cloned workbook — %d records moved to UNKNOWN",
                sheet_name, len(rec_list),
            )
            unknown_records.extend(rec_list)
            continue

        ws = wb[sheet_name]
        sc = schema[sheet_name]

        # Verify the IGNORE service row is intact.
        if "Action" in sc.col_index:
            action_col = sc.col_index["Action"] + 1
            row11_action = ws.cell(row=11, column=action_col).value
            if str(row11_action or "").strip().upper() != "IGNORE":
                raise RuntimeError(
                    f"Sheet {sheet_name!r}: row 11 Action is {row11_action!r}, "
                    "expected 'IGNORE'. The template may be corrupted."
                )

        # Write data directly into the template's blank example rows starting at
        # _FIRST_DATA.  No insert_rows needed — the template already has ~1000
        # blank rows between row 12 and LAST ROW.
        for i, rec in enumerate(rec_list):
            row_idx = _FIRST_DATA + i
            if row_idx >= sc.last_row_index:
                logger.warning(
                    "Sheet %s: row %d would reach LAST ROW at %d — "
                    "%d of %d records skipped",
                    sheet_name, row_idx, sc.last_row_index,
                    len(rec_list) - i, len(rec_list),
                )
                break
            row_data: dict[str, object] = {}

            # Mandatory fields — always written for every equipment type
            raw_area = rec.get("PARENT_AREA")
            if not raw_area or str(raw_area).strip() in ('', 'nan', 'None'):
                parent_area_val = default_area
            else:
                parent_area_val = str(raw_area).strip()
            parent_areas.add(parent_area_val)

            user_tag_raw = str(rec.get("USER_TAG") or "").strip()

            desc_raw = str(rec.get("DESCRIPTION") or rec.get("DESCRIPTION_RU") or "")

            _set_if_exists(row_data, sc, "Action",            "NEW")
            _set_if_exists(row_data, sc, "ObjectName",        f"{sheet_name}   {user_tag_raw}")
            _set_if_exists(row_data, sc, "CpUserTag",         _truncate(user_tag_raw.replace('-', ''), 20))
            _set_if_exists(row_data, sc, "CpItemDescription", user_tag_raw)
            _set_if_exists(row_data, sc, "ParentArea",        parent_area_val)
            _set_if_exists(row_data, sc, "CpNumberItems",     int(rec.get("QUANTITY") or 1))
            _set_if_exists(row_data, sc, "CpApplication",     _status(rec.get("OPERATION_STATUS")))

            # Generic field mapping via UNIVERSAL_MAP
            for field_name, slot_name in UNIVERSAL_MAP.items():
                if slot_name in MANDATORY_SLOTS:
                    continue
                if slot_name not in sc.col_index:
                    continue

                xlsx_col = RECORD_COLUMN_MAP.get(field_name, field_name.upper())
                raw_value = rec.get(xlsx_col)

                # Boolean/status slots handled specially (skip unit conversion)
                if slot_name == 'CpVFD':
                    row_data[slot_name] = _bool_to_yn(raw_value)
                    continue
                elif slot_name == 'CpExplosionProof':
                    row_data[slot_name] = _bool_to_yn(raw_value)
                    continue

                if _is_nan(raw_value):
                    continue

                source_unit = _get_source_unit(rec, field_name)
                col_idx = sc.col_index[slot_name]
                target_unit = sc.units[col_idx] if col_idx < len(sc.units) else ""

                converted = convert(raw_value, source_unit, target_unit)
                if converted is not None:
                    row_data[slot_name] = converted

            # Write the assembled row into the worksheet (skip col A — ACCE-reserved).
            for slot_name, value in row_data.items():
                col_idx = sc.col_index[slot_name]
                if col_idx + 1 > 1:
                    ws.cell(row=row_idx, column=col_idx + 1).value = value

        by_sheet[sheet_name] = len(rec_list)
        logger.info("Wrote %d records to sheet %s", len(rec_list), sheet_name)

    # ── Step 5: populate AREAS sheet ─────────────────────────────────────────
    _write_area(wb, default_area, parent_areas)

    # ── Step 6: write review sheets ──────────────────────────────────────────
    write_unknown_sheet(wb, unknown_records, skipped_invalid)

    # ── Step 7: save ─────────────────────────────────────────────────────────
    wb.save(output_path)
    fix_inline_strings(output_path)
    logger.info("Saved ACCE export to %s", output_path)

    exported = sum(by_sheet.values())
    return ExportSummary(
        total_records=len(records),
        exported=exported,
        skipped_invalid=len(skipped_invalid),
        unknown_symbol=len(unknown_records),
        used_default_type=used_default_type,
        by_sheet=by_sheet,
    )


# ── Private helpers ───────────────────────────────────────────────────────────

def fix_inline_strings(xlsx_path: str) -> None:
    """Convert openpyxl's t='inlineStr' cells to t='s' (sharedStrings).

    openpyxl writes every string cell as inlineStr and omits sharedStrings.xml.
    ACCE cannot read inlineStr and treats those cells as empty, which causes
    the IGNORE-row CpUserTag to appear blank — duplicating the LAST ROW blank
    and triggering "Duplicate user tag number found" on every equipment sheet.
    """
    tmp_path = xlsx_path + '.tmp'
    shutil.copy2(xlsx_path, tmp_path)

    shared_strings: list[str] = []
    ss_index: dict[str, int]  = {}

    def get_ss_index(s: str) -> int:
        if s not in ss_index:
            ss_index[s] = len(shared_strings)
            shared_strings.append(s)
        return ss_index[s]

    # Pass 1: collect all inlineStr text values so the shared-string table is
    # complete before any sheet is rewritten.
    sheet_xmls: dict[str, str] = {}
    with zipfile.ZipFile(tmp_path, 'r') as zin:
        for name in zin.namelist():
            if re.match(r'xl/worksheets/sheet\d+\.xml$', name):
                xml = zin.read(name).decode('utf-8')
                for m in re.finditer(
                        r'<c([^>]*)t="inlineStr"([^>]*)>'
                        r'<is><t[^>]*>([^<]*)</t></is></c>', xml):
                    get_ss_index(m.group(3))
                sheet_xmls[name] = xml

    # Pass 2: rewrite sheets and reassemble the zip in one pass to avoid
    # duplicate entries for [Content_Types].xml and workbook.xml.rels.
    def replace_inline(xml: str) -> str:
        def repl(m: re.Match) -> str:
            attrs = (m.group(1) + m.group(2)).replace('t="inlineStr"', '').strip()
            idx   = get_ss_index(m.group(3))
            return f'<c {attrs} t="s"><v>{idx}</v></c>'
        return re.sub(
            r'<c([^>]*)t="inlineStr"([^>]*)>'
            r'<is><t[^>]*>([^<]*)</t></is></c>',
            repl, xml)

    ss_items = ''.join(f'<si><t>{s}</t></si>' for s in shared_strings)
    ss_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main" '
        f'count="{len(shared_strings)}" '
        f'uniqueCount="{len(shared_strings)}">'
        f'{ss_items}</sst>'
    )

    with zipfile.ZipFile(tmp_path, 'r') as zin:
        needs_ss = 'xl/sharedStrings.xml' not in zin.namelist()
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                if name in sheet_xmls:
                    data = replace_inline(sheet_xmls[name]).encode('utf-8')
                elif name == 'xl/sharedStrings.xml':
                    data = ss_xml.encode('utf-8')
                elif name == '[Content_Types].xml' and needs_ss:
                    ct = zin.read(name).decode('utf-8')
                    ct = ct.replace(
                        '</Types>',
                        '<Override PartName="/xl/sharedStrings.xml" '
                        'ContentType="application/vnd.openxmlformats-'
                        'officedocument.spreadsheetml.sharedStrings+xml"/>'
                        '</Types>')
                    data = ct.encode('utf-8')
                elif name == 'xl/_rels/workbook.xml.rels' and needs_ss:
                    wr = zin.read(name).decode('utf-8')
                    wr = wr.replace(
                        '</Relationships>',
                        '<Relationship Id="rIdSS" '
                        'Type="http://schemas.openxmlformats.org/'
                        'officeDocument/2006/relationships/sharedStrings" '
                        'Target="sharedStrings.xml"/>'
                        '</Relationships>')
                    data = wr.encode('utf-8')
                else:
                    data = zin.read(name)
                zout.writestr(name, data)

            if needs_ss:
                zout.writestr('xl/sharedStrings.xml', ss_xml.encode('utf-8'))

    os.remove(tmp_path)


def _write_area(wb, default_area: str, parent_areas: set[str]) -> None:
    """Write AREAS sheet with NEW rows for the default area and each unique parent area."""
    if "AREAS" not in wb.sheetnames:
        logger.warning("AREAS sheet not found in workbook — skipping area population")
        return
    ws = wb["AREAS"]

    # Find the LAST ROW marker in AREAS so we don't unmerge the
    # GENERAL INSTRUCTIONS block that sits below it.
    areas_last_row = ws.max_row + 1
    for _r in range(1, ws.max_row + 2):
        _v = ws.cell(row=_r, column=1).value
        if _v is not None and str(_v).strip().upper() == "LAST ROW":
            areas_last_row = _r
            break

    for merge_range in list(ws.merged_cells.ranges):
        if (merge_range.max_row >= 11  # rows from 11 are written by this function
                and merge_range.min_row < areas_last_row):
            ws.unmerge_cells(str(merge_range))

    # Build 1-based column map from the SlotNames row (row 5).
    areas_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=5, column=c).value
        if v and str(v).strip() not in ("SlotNames", "LAST COLUMN"):
            areas_col[str(v).strip()] = c

    def _set_row(row: int, slot: str, value: object) -> None:
        col = areas_col.get(slot)
        if col and col > 1:   # never write to column A — ACCE-reserved
            if isinstance(ws.cell(row=row, column=col), _MergedCell):
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                        ws.unmerge_cells(str(mr))
                        break
            ws.cell(row=row, column=col).value = value

    # Build the ordered list: default_area first, then sorted extra areas.
    extra_areas = sorted(a for a in parent_areas if a != default_area)
    area_sequence = [default_area] + extra_areas   # rows 11, 12, 13, …

    # Write NEW rows starting at row 11. Column A is never touched.
    for i, area in enumerate(area_sequence):
        row_idx = 11 + i
        _set_row(row_idx, "Action",             "NEW")
        _set_row(row_idx, "ArAreaSubAreaTitle", area)
        _set_row(row_idx, "ArReportGroupTitle", default_area)
        _set_row(row_idx, "ObjectName",         f"__AREA__{area} ")
        _set_row(row_idx, "ArAreaType",         "GRADE")

    logger.debug(
        "AREAS: %d NEW rows (default=%r + %d extra)",
        len(area_sequence), default_area, len(extra_areas),
    )


def _set_if_exists(
    row_data: dict[str, object],
    sc: SheetSchema,
    slot_name: str,
    value: object,
) -> None:
    if slot_name in sc.col_index:
        row_data[slot_name] = value


def _truncate(value: str, max_len: int) -> str:
    """Sanitise USER_TAG for ACCE: replace invalid chars, truncate to max_len."""
    cleaned = (
        value.replace(" ", "_")
             .replace("/", "-")
             .replace("@", "at")
    )
    if len(cleaned) > max_len:
        logger.warning(
            "USER_TAG %r truncated to %d characters → %r",
            value, max_len, cleaned[:max_len],
        )
        cleaned = cleaned[:max_len]
    return cleaned


def _bool_to_yn(val) -> str:
    """Convert parser boolean/string value to ACCE YES/NO string."""
    s = str(val or '').upper()
    if s in ('TRUE', 'YES', 'ДА', '1'):
        return 'YES'
    if s in ('FALSE', 'NO', 'НЕТ', '0'):
        return 'NO'
    return ''


def _status(val) -> str:
    """Map parser operation_status to ACCE CpApplication value."""
    s = str(val or '').upper()
    return 'SPARE' if any(x in s for x in ('SPAR', 'STANDBY', 'РЕЗЕРВ')) else 'CONT'


def _get_source_unit(record: dict, field_name: str) -> str:
    """Return the source unit string for a given field from the parser record."""
    if field_name in ("pressure", "pressure_fan", "pressure_steam"):
        unit = record.get("PRESSURE_UNIT")
        return str(unit).strip() if unit and not _is_nan(unit) else "MPa"
    if field_name in ("flow_rate", "flow_rate_fan"):
        unit = record.get("FLOW_RATE_UNIT")
        return str(unit).strip() if unit and not _is_nan(unit) else "m3/h"
    return _SOURCE_UNITS.get(field_name, "")


def _is_nan(value: object) -> bool:
    """Return True if value represents a missing/null value."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and value.strip().lower() in ("nan", "none", ""):
        return True
    return False


