"""Reads master_template.xlsx and extracts sheet schemas."""

from __future__ import annotations

import logging
from dataclasses import dataclass, field

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class SheetSchema:
    sheet_name: str
    slot_names: list[str]       # row 4: CpUserTag, ParentArea, ...
    human_names: list[str]      # row 5
    units: list[str]            # row 9
    col_index: dict[str, int]   # slot_name → 0-based column index
    header_rows: int = 9        # rows 1–9 to preserve as-is
    last_row_index: int = 0     # 1-based row index of "LAST ROW"
    template_max_row: int = 0   # max_row of the sheet in the original template
    template_wb: object = field(default=None, repr=False)  # openpyxl.Workbook


_SLOT_ROW   = 5   # row containing slot names (first cell = 'SlotNames')
_NAME_ROW   = 6   # row containing human-readable names
_UNIT_ROW   = 10  # row containing unit strings (first cell = 'Units')
_FIRST_DATA = 12  # first data row; row 11 is the IGNORE service row — never overwrite it

# Labels that appear as the first cell in their respective rows — not slots.
_ROW_LABELS = frozenset({"SlotNames", "Name", "Min", "Max", "Default", "Units"})
# Column marker that signals end of slot columns — not a writable slot.
_LAST_COL_MARKER = "LAST COLUMN"


def _is_equipment_sheet(ws) -> bool:
    """Return True for individual-equipment sheets (not bulk material sheets).

    Equipment sheets have both CpItemDescription and ParentArea in row 5;
    bulk sheets (Pipe, Civil, Electrical, etc.) share CpUserTag but not these.
    """
    row5 = {ws.cell(row=_SLOT_ROW, column=c).value for c in range(1, ws.max_column + 1)}
    return "CpItemDescription" in row5 and "ParentArea" in row5


def read_template(template_path: str) -> dict[str, SheetSchema]:
    """Open template, detect equipment sheets by slot content, return schema dict."""
    wb = openpyxl.load_workbook(template_path, keep_vba=False, data_only=True)
    schemas: dict[str, SheetSchema] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if not _is_equipment_sheet(ws):
            continue

        slot_names = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[_SLOT_ROW]
        ]
        human_names = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[_NAME_ROW]
        ]
        units = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in ws[_UNIT_ROW]
        ]

        # Skip the row-label cell (col 1) and the LAST COLUMN sentinel.
        col_index: dict[str, int] = {
            name: idx
            for idx, name in enumerate(slot_names)
            if name and name not in _ROW_LABELS and name != _LAST_COL_MARKER
        }

        # Find the row whose first cell contains "LAST ROW"
        last_row_index = 0
        for row_idx in range(1, ws.max_row + 2):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value is not None and str(cell_value).strip().upper() == "LAST ROW":
                last_row_index = row_idx
                break

        if last_row_index == 0:
            logger.warning(
                "Sheet %s: 'LAST ROW' marker not found; defaulting to max_row + 1",
                sheet_name,
            )
            last_row_index = ws.max_row + 1

        schemas[sheet_name] = SheetSchema(
            sheet_name=sheet_name,
            slot_names=slot_names,
            human_names=human_names,
            units=units,
            col_index=col_index,
            last_row_index=last_row_index,
            template_max_row=ws.max_row,
            template_wb=wb,
        )

        logger.debug(
            "Loaded schema for %s: %d slots, LAST ROW at %d",
            sheet_name,
            len(col_index),
            last_row_index,
        )

    logger.info("Template loaded: %d equipment sheets from %s", len(schemas), template_path)
    return schemas
