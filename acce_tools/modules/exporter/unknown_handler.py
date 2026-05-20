"""Handles UNKNOWN and skipped records — writes them to a separate sheet
in the output file for reporter review."""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_COLS = [
    "USER_TAG",
    "RAW_TAG",
    "DESCRIPTION",
    "ACCE_ITEM_SYMBOL",
    "ACCE_ITEM_TYPE",
    "IS_VALID",
    "ERRORS",
    "WARNINGS",
    "SOURCE_FILE",
    "SOURCE_ROW",
]

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_HEADER_FONT = Font(bold=True)


def write_unknown_sheet(
    wb: openpyxl.Workbook,
    unknown_records: list[dict],
    skipped_records: list[dict],
) -> None:
    """Create UNKNOWN and SKIPPED sheets in *wb* with review information."""
    _write_review_sheet(wb, "UNKNOWN", unknown_records, "Total UNKNOWN")
    _write_review_sheet(wb, "SKIPPED", skipped_records, "Total SKIPPED")


def _write_review_sheet(
    wb: openpyxl.Workbook,
    title: str,
    records: list[dict],
    summary_label: str,
) -> None:
    if title in wb.sheetnames:
        del wb[title]

    ws = wb.create_sheet(title=title)

    # Header row
    ws.append(_COLS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"

    # Data rows
    for rec in records:
        row = [_safe_str(rec.get(col)) for col in _COLS]
        ws.append(row)

    # Summary row
    ws.append([f"{summary_label}: {len(records)}"])

    # Auto column widths (based on content)
    for col_idx, col_name in enumerate(_COLS, start=1):
        col_letter = get_column_letter(col_idx)
        max_width = max(
            (len(str(cell.value or "")) for cell in ws[col_letter]),
            default=len(col_name),
        )
        ws.column_dimensions[col_letter].width = min(max_width + 2, 50)


def _safe_str(value: object) -> str:
    if value is None:
        return ""
    s = str(value)
    return "" if s.lower() in ("nan", "none") else s
