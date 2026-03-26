"""
parser_xlsx.py
--------------
Parser for equipment list Excel files in the format used by this project.

Confirmed against: Пример_4.xlsx
Sheet used:        'Лист1'  (the main equipment list — other sheets are price
                             reference tables and are intentionally skipped)

Column layout (confirmed from file):
  A  (col 1)  № п/п               — sequential number
  B  (col 2)  ТИТУЛ               — area / contract section code
  C  (col 3)  ТАГОВЫЙ НОМЕР       — equipment tag
  D  (col 4)  НАИМЕНОВАНИЕ        — description (bilingual: Russian / English)
  E  (col 5)  код ед.расценки     — price/type code (e.g. 6831, 6880.12)
  F  (col 6)  ТЕХ. ХАРАКТЕРИСТИКИ — technical characteristics string
  G  (col 7)  КОЛ.                — quantity
  H  (col 8)  ВЕС ЕДИНИЦЫ (кг)   — unit weight kg
  I  (col 9)  (merged/empty)
  J  (col 10) Всего масса, кг     — total mass kg
  K  (col 11) Всего масса, ТОННЫ  — total mass tonnes

Header structure:
  Row 1: column names  (row 2 is merged continuation — empty values)
  Rows 3+: data

Usage:
    from modules.parser_xlsx import parse_xlsx
    records, parse_warnings = parse_xlsx("data/samples/Пример_4.xlsx")
"""

import re
import openpyxl
from modules.schema import ACCERecord

# ─────────────────────────────────────────────────────────────────────────────
# PRICE CODE → ACCE equipment model mapping
# Source: cross-reference of price codes in the file with ACCE documentation
# Extend this dict as more equipment types appear in future files.
# ─────────────────────────────────────────────────────────────────────────────
PRICE_CODE_TO_ACCE = {
    "6811":    "GEN BLOCK",       # Equipment supplied as single block unit
    "6820":    "GEN STATIC",      # Vertical/horizontal static equipment (vessels, tanks, valves)
    "6831":    "GEN FIRED",       # Fired equipment (furnaces, incinerators, boilers)
    "6852":    "GEN ROTATING",    # Rotating equipment (pumps, fans, compressors)
    "6870":    "GEN AIRCOOLER",   # Air coolers
    "6880.12": "GEN PACKAGE",     # Package units / systems
    "6880.13": "GEN CRANE",       # Overhead cranes, hoists
    "2521-01": "GEN TANK",        # Storage tanks (rolled construction)
}

# ─────────────────────────────────────────────────────────────────────────────
# Column index constants  (1-based, matching confirmed file layout)
# ─────────────────────────────────────────────────────────────────────────────
COL_SEQ         = 1
COL_TITLE       = 2
COL_TAG         = 3
COL_DESC        = 4
COL_PRICE_CODE  = 5
COL_TECH        = 6
COL_QTY         = 7
COL_WEIGHT_UNIT = 8
COL_WEIGHT_TOT  = 10
COL_WEIGHT_TOT_T= 11

HEADER_ROW      = 1   # row containing column names
DATA_START_ROW  = 3   # first row with actual equipment data


# ─────────────────────────────────────────────────────────────────────────────
# Low-level helpers
# ─────────────────────────────────────────────────────────────────────────────

def _build_merge_map(sheet) -> dict:
    """
    openpyxl returns None for cells inside a merged range (except top-left).
    This builds a (row, col) → value lookup so we never lose merged values.
    """
    merge_map = {}
    for merged_range in sheet.merged_cells.ranges:
        top_val = sheet.cell(
            row=merged_range.min_row,
            column=merged_range.min_col
        ).value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(r, c)] = top_val
    return merge_map


def _cell(sheet, row: int, col: int, merge_map: dict):
    """Get cell value, falling back to merged-cell map if the cell is None."""
    val = sheet.cell(row=row, column=col).value
    if val is None:
        val = merge_map.get((row, col))
    return val


def _clean_str(value) -> str:
    """Strip and normalise a string value."""
    if value is None:
        return ""
    # Normalise internal newlines and extra spaces
    return re.sub(r'\s+', ' ', str(value).strip())


def _clean_float(value) -> float | None:
    """Convert a cell value to float, handling None, empty, and text."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "-", "n/a", "tbd", "tbc", "#REF!"):
        return None
    try:
        return float(s)
    except ValueError:
        # Try extracting first number from a string like "234,800"
        m = re.search(r'-?\d[\d,\.]*', s)
        if m:
            try:
                return float(m.group().replace(',', ''))
            except ValueError:
                return None
    return None


def _clean_int(value) -> int | None:
    """Convert a cell value to int."""
    f = _clean_float(value)
    return int(f) if f is not None else None


def _split_bilingual(text: str) -> tuple[str, str]:
    """
    Split a bilingual description like:
        'Насос для оборотной воды / Circulating Water Pump'
        'Инсинератор/ INCINERATOR'
    Returns (russian_part, english_part).
    """
    if '/' in text:
        parts = text.split('/', 1)
        return parts[0].strip(), parts[1].strip()
    return text.strip(), ""


# ─────────────────────────────────────────────────────────────────────────────
# Technical characteristics parser
# Handles all patterns confirmed in the sample file:
#   Q=5,000-5,500m3/h P=2,500-3,000Pa
#   Q=4000kW  P=0.8MPa
#   Q=160m3/h  P=0.8MPa
#   Φ3.8×25m
#   DN500
#   V=6m3
#   T=5t H=6m
# ─────────────────────────────────────────────────────────────────────────────

_UNITS = r'(m3/h|m3|kW|MPa|kPa|Pa|t/h|kg/h|t|m)?'

def _extract_param(text: str, key: str) -> tuple[float | None, str]:
    """
    Extract a named parameter like Q=160m3/h or P=0.8MPa.
    Returns (value, unit). For ranges (5,000-5,500) returns the lower bound.
    """
    pattern = (
        rf'{re.escape(key)}\s*=\s*'
        rf'([0-9][0-9,\.]*)'     # first number (possibly with commas)
        rf'(?:\s*[-~]\s*[0-9][0-9,\.]*)?'  # optional range upper bound
        rf'\s*' + _UNITS
    )
    m = re.search(pattern, text, re.IGNORECASE)
    if not m:
        return None, ""
    raw = m.group(1).replace(',', '')
    unit = m.group(2) or ""
    try:
        return float(raw), unit
    except ValueError:
        return None, ""


def _extract_dn(text: str) -> int | None:
    m = re.search(r'DN\s*(\d+)', text, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _extract_phi_dimensions(text: str) -> tuple[float | None, float | None]:
    """Parse Φ3.8×25m → (diameter_m=3.8, length_m=25.0)."""
    m = re.search(r'[ΦΦ]?\s*([\d.]+)\s*[×x]\s*([\d.]+)', text)
    if m:
        return float(m.group(1)), float(m.group(2))
    return None, None


def _extract_lift(text: str) -> float | None:
    """Parse T=5t → 5.0 (crane/hoist lifting capacity in tonnes)."""
    m = re.search(r'T\s*=\s*([\d.]+)\s*t', text)
    return float(m.group(1)) if m else None


def _parse_tech_characteristics(raw: str) -> dict:
    """
    Parse the ТЕХ. ХАРАКТЕРИСТИКИ string into structured fields.
    Returns a dict of extracted values (None for anything not found).
    """
    if not raw:
        return {}

    result = {}

    q_val, q_unit = _extract_param(raw, 'Q')
    p_val, p_unit = _extract_param(raw, 'P')
    v_val, v_unit = _extract_param(raw, 'V')
    dn             = _extract_dn(raw)
    dia, length    = _extract_phi_dimensions(raw)
    lift           = _extract_lift(raw)

    # Q can mean flow rate (m3/h) or thermal duty (kW) — keep both separately
    if q_val is not None:
        if q_unit and q_unit.lower() == 'kw':
            result['capacity_kw']    = q_val
        else:
            result['flow_rate']      = q_val
            result['flow_rate_unit'] = q_unit

    if p_val is not None:
        result['pressure']      = p_val
        result['pressure_unit'] = p_unit

    if v_val is not None:
        result['volume']      = v_val
        result['volume_unit'] = v_unit

    if dn is not None:
        result['dn_mm'] = dn

    if dia is not None:
        result['diameter_m'] = dia
    if length is not None:
        result['length_m'] = length

    if lift is not None:
        result['lift_capacity_t'] = lift

    return result


# ─────────────────────────────────────────────────────────────────────────────
# Tag sanitiser — make raw tags safe for ACCE user_tag field
# Rules from ACCE docs: max 20 chars, no spaces, no special characters
# ─────────────────────────────────────────────────────────────────────────────

_FORBIDDEN = re.compile(r'[!@#$%&*()?/{}[\]*=<>,`~^:;|"\'\\+\s]')

def _make_user_tag(raw_tag: str, seq: int) -> str:
    """
    Convert a raw equipment tag to a valid ACCE user tag.
    Strips forbidden characters, truncates to 20 chars.
    Falls back to SEQ-{n} if the result is empty.
    """
    safe = _FORBIDDEN.sub('-', raw_tag.strip())
    safe = re.sub(r'-{2,}', '-', safe).strip('-')
    safe = safe[:20]
    return safe if safe else f"SEQ-{seq:04d}"


# ─────────────────────────────────────────────────────────────────────────────
# Main parser
# ─────────────────────────────────────────────────────────────────────────────

DATA_SHEET = 'Лист1'   # the confirmed name of the equipment list sheet

def parse_xlsx(filepath: str) -> tuple[list[ACCERecord], list[dict]]:
    """
    Parse an equipment list Excel file.

    Args:
        filepath: path to the .xlsx file

    Returns:
        records:       list[ACCERecord] — one per equipment row
        parse_warnings: list[dict]      — rows skipped or partially parsed
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    if DATA_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Expected sheet '{DATA_SHEET}' not found. "
            f"Available sheets: {wb.sheetnames}"
        )

    sheet = wb[DATA_SHEET]
    merge_map = _build_merge_map(sheet)
    records: list[ACCERecord] = []
    parse_warnings: list[dict] = []

    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):

        # ── Skip completely empty rows ────────────────────────────────────────
        row_vals = [_cell(sheet, row_idx, c, merge_map) for c in range(1, 12)]
        if all(v is None for v in row_vals):
            continue

        # ── Read raw cell values ──────────────────────────────────────────────
        seq        = _clean_int (_cell(sheet, row_idx, COL_SEQ,         merge_map))
        title      = _clean_str (_cell(sheet, row_idx, COL_TITLE,       merge_map))
        tag        = _clean_str (_cell(sheet, row_idx, COL_TAG,         merge_map))
        desc_raw   = _clean_str (_cell(sheet, row_idx, COL_DESC,        merge_map))
        price_code = _clean_str (_cell(sheet, row_idx, COL_PRICE_CODE,  merge_map))
        tech_raw   = _clean_str (_cell(sheet, row_idx, COL_TECH,        merge_map))
        qty        = _clean_int (_cell(sheet, row_idx, COL_QTY,         merge_map))
        w_unit     = _clean_float(_cell(sheet, row_idx, COL_WEIGHT_UNIT,merge_map))
        w_total    = _clean_float(_cell(sheet, row_idx, COL_WEIGHT_TOT, merge_map))
        w_total_t  = _clean_float(_cell(sheet, row_idx, COL_WEIGHT_TOT_T,merge_map))

        # ── Skip summary / total rows (no tag, no sequence number) ────────────
        if not tag and not seq:
            parse_warnings.append({
                "sheet": DATA_SHEET,
                "row": row_idx,
                "issue": "No tag and no sequence number — likely a totals row, skipped",
            })
            continue

        # ── Warn about rows with missing critical fields ───────────────────────
        row_warnings = []
        if not tag:
            row_warnings.append("Missing equipment tag")
        if not desc_raw:
            row_warnings.append("Missing description")
        if not price_code:
            row_warnings.append("Missing price/type code")

        # ── Parse bilingual description ───────────────────────────────────────
        desc_ru, desc_en = _split_bilingual(desc_raw)

        # ── Parse technical characteristics ───────────────────────────────────
        tech = _parse_tech_characteristics(tech_raw)

        # ── Build ACCERecord ──────────────────────────────────────────────────
        record = ACCERecord(
            # Source tracing
            source_file  = filepath,
            source_sheet = DATA_SHEET,
            source_row   = row_idx,

            # Identity
            seq_number      = seq or 0,
            raw_tag         = tag,
            description_ru  = desc_ru,
            description_en  = desc_en,
            price_code      = price_code,
            quantity        = qty or 1,
            weight_unit_kg  = w_unit,
            weight_total_kg = w_total,
            weight_total_t  = w_total_t,
            title           = title,

            # Technical characteristics
            tech_raw        = tech_raw,
            flow_rate       = tech.get('flow_rate'),
            flow_rate_unit  = tech.get('flow_rate_unit', ''),
            pressure        = tech.get('pressure'),
            pressure_unit   = tech.get('pressure_unit', ''),
            volume          = tech.get('volume'),
            volume_unit     = tech.get('volume_unit', ''),
            capacity_kw     = tech.get('capacity_kw'),
            diameter_m      = tech.get('diameter_m'),
            length_m        = tech.get('length_m'),
            dn_mm           = tech.get('dn_mm'),
            lift_capacity_t = tech.get('lift_capacity_t'),

            # ACCE control fields (pre-filled, finalized by Standardiser)
            action         = "NEW",
            user_tag       = _make_user_tag(tag, seq or row_idx),
            parent_area    = title,
            acce_equip_type= PRICE_CODE_TO_ACCE.get(price_code, ""),

            # Validation state
            is_valid = len(row_warnings) == 0,
            warnings = row_warnings,
        )

        records.append(record)

        if row_warnings:
            parse_warnings.append({
                "sheet": DATA_SHEET,
                "row":   row_idx,
                "tag":   tag,
                "issue": "; ".join(row_warnings),
            })

    return records, parse_warnings
