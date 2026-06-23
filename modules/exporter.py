import openpyxl
from openpyxl.styles import Font, Alignment
from typing import List, Dict
import logging
import os
from dataclasses import dataclass

from .schema import ACCERecord

logger = logging.getLogger(__name__)

SHEET_NAME_MAP = {
    'CP': 'CP',  # Pumps
    'FN': 'FN',  # Fans
    'GC': 'GC',  # Compressors
    'VT': 'VT',  # Vessels/Tanks
    'HT': 'HT',  # Horizontal Tanks
    'HE': 'HE',  # Heat Exchangers
    'FU': 'FU',  # Furnaces
    'STB': 'STB',  # Boilers
    'CE': 'CE',  # Cranes
    'HO': 'HO',  # Hoists
    'FLR': 'FLR',  # Flares
    'STK': 'STK',  # Stacks
    'DC': 'DC',  # Dust Collectors
}


@dataclass
class ExportResult:
    exported_count: int
    skipped_count: int
    by_type: Dict[str, int]
    output_path: str
    warnings: List[str]


def _ensure_unique_tags(records: List[ACCERecord]) -> List[str]:
    seen_tags = {}
    warnings = []

    tags_list = [rec.user_tag for rec in records if rec.is_valid and rec.user_tag]

    for i, tag in enumerate(tags_list):
        if not tag:
            continue

        base_tag = tag
        counter = 1

        while tag in seen_tags:
            suffix = f"-{chr(64 + counter)}"  # -A, -B...
            max_base_len = 20 - len(suffix)
            new_tag = base_tag[:max_base_len] + suffix
            tag = new_tag
            counter += 1

        seen_tags[tag] = True
        if tag != records[i].user_tag:
            warnings.append(f"Duplicate tag '{records[i].user_tag}' renamed to '{tag}'")
            records[i].user_tag = tag

    return warnings


def export_to_acce(
        records: List[ACCERecord],
        template_path: str,
        output_path: str,
        project_name: str = ""
) -> ExportResult:

    logger.info(f"Starting export using template: {template_path}")

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found at: {template_path}. Please provide a valid ACCE template.")

    valid_records = [r for r in records if r.is_valid]
    skipped_count = len(records) - len(valid_records)

    if not valid_records:
        logger.warning("No valid records to export.")

    tag_warnings = _ensure_unique_tags(valid_records)

    areas_set = set()
    for rec in valid_records:
        area = rec.parent_area.strip() if rec.parent_area else "DEFAULT"
        areas_set.add(area)

    grouped_records: Dict[str, List[ACCERecord]] = {}
    for rec in valid_records:
        symbol = rec.acce_item_symbol
        if not symbol:
            continue
        if symbol not in grouped_records:
            grouped_records[symbol] = []
        grouped_records[symbol].append(rec)

    sorted_symbols = sorted(grouped_records.keys())

    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        raise IOError(f"Cannot load template from {template_path}: {e}")

    if 'AREAS' in wb.sheetnames:
        ws_areas = wb['AREAS']
        last_row_idx = None
        for r in range(1, ws_areas.max_row + 1):
            if str(ws_areas.cell(r, 1).value).upper() == "LAST ROW":
                last_row_idx = r
                break

        if not last_row_idx:
            last_row_idx = ws_areas.max_row + 1
            ws_areas.cell(last_row_idx, 1, value="LAST ROW")

        current_insert_row = last_row_idx
        for area_name in sorted(areas_set):
            ws_areas.insert_rows(current_insert_row, amount=1)

            # TODO: Проверить названия колонок в реальном шаблоне AREAS.
            # Сейчас предполагаем: A=ACTION, B=AREA_NAME, C=REPORT_GROUP, D=AREA_TYPE
            ws_areas.cell(current_insert_row, 1, value="NEW")  # ACTION
            ws_areas.cell(current_insert_row, 2, value=area_name)  # AREA_NAME
            ws_areas.cell(current_insert_row, 3, value=area_name)  # REPORT_GROUP
            ws_areas.cell(current_insert_row, 4, value="PROCESS")  # AREA_TYPE

            current_insert_row += 1
    else:
        logger.error("Sheet 'AREAS' not found in template!")

    by_type_counts = {}

    for symbol in sorted_symbols:
        sheet_name = SHEET_NAME_MAP.get(symbol, symbol)

        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in template. Skipping type {symbol}.")
            continue

        ws_eq = wb[sheet_name]

        last_row_idx = None
        for r in range(1, ws_eq.max_row + 1):
            if str(ws_eq.cell(r, 1).value).upper() == "LAST ROW":
                last_row_idx = r
                break

        if not last_row_idx:
            last_row_idx = ws_eq.max_row + 1
            ws_eq.cell(last_row_idx, 1, value="LAST ROW")

        current_insert_row = last_row_idx
        records_for_type = grouped_records[symbol]

        for rec in records_for_type:
            ws_eq.insert_rows(current_insert_row, amount=1)

            # TODO: Убедиться, что в реальном шаблоне колонки A-F имеют именно такие значения или смысл.
            ws_eq.cell(current_insert_row, 1, value=rec.action)  # A: ACTION
            ws_eq.cell(current_insert_row, 2, value=rec.acce_item_symbol)  # B: ITEM_SYMBOL
            ws_eq.cell(current_insert_row, 3, value=rec.acce_item_type)  # C: ITEM_TYPE
            ws_eq.cell(current_insert_row, 4, value=rec.user_tag)  # D: USER_TAG

            area_val = rec.parent_area.strip() if rec.parent_area else "DEFAULT"
            ws_eq.cell(current_insert_row, 5, value=area_val)  # E: PARENT_AREA

            desc = rec.description_ru[:60] if rec.description_ru else ""
            ws_eq.cell(current_insert_row, 6, value=desc)  # F: DESCRIPTION

            col_idx = 7

            press_kpa = rec.pressure * 1000 if rec.pressure else None

            if symbol == 'CP':  # Насосы
                # TODO: Проверить порядок колонок в реальном шаблоне CP
                ws_eq.cell(current_insert_row, col_idx, value=rec.design_temperature)  # G
                ws_eq.cell(current_insert_row, col_idx + 1, value=press_kpa)  # H
                ws_eq.cell(current_insert_row, col_idx + 2, value=rec.motor_power_kw)  # I

            elif symbol == 'FN':  # Вентиляторы
                # TODO: Проверить порядок колонок в реальном шаблоне FN
                fn_press = rec.pressure
                if rec.pressure_unit == 'МПа':
                    fn_press = rec.pressure * 1_000_000

                ws_eq.cell(current_insert_row, col_idx, value=rec.flow_rate)  # G
                ws_eq.cell(current_insert_row, col_idx + 1, value=fn_press)  # H
                ws_eq.cell(current_insert_row, col_idx + 2, value=rec.motor_power_kw)  # I

            elif symbol == 'VT':  # Емкости
                # TODO: Проверить порядок колонок в реальном шаблоне VT
                ws_eq.cell(current_insert_row, col_idx, value=press_kpa)  # G
                ws_eq.cell(current_insert_row, col_idx + 1, value=rec.design_temperature)  # H
                ws_eq.cell(current_insert_row, col_idx + 2, value=rec.volume)  # I
                ws_eq.cell(current_insert_row, col_idx + 3, value=rec.diameter_m)  # J
                ws_eq.cell(current_insert_row, col_idx + 4, value=rec.material)  # K

            elif symbol == 'STB':  # Котлы
                # TODO: Проверить порядок колонок в реальном шаблоне STB
                ws_eq.cell(current_insert_row, col_idx, value=rec.capacity_kw)  # G
                ws_eq.cell(current_insert_row, col_idx + 1, value=press_kpa)  # H

            current_insert_row += 1

        by_type_counts[symbol] = len(records_for_type)

    if 'VALIDATION' in wb.sheetnames:
        ws_val = wb['VALIDATION']
        for row in ws_val.iter_rows(min_row=2, max_col=5, values_only=False):
            for cell in row:
                cell.value = None
    else:
        ws_val = wb.create_sheet("VALIDATION")
        val_headers = ["TAG", "TYPE", "STATUS", "ERRORS", "WARNINGS"]
        for c, h in enumerate(val_headers, 1):
            ws_val.cell(1, c, value=h).font = Font(bold=True)

    row_val = 2
    for rec in records:
        status = "VALID" if rec.is_valid else "INVALID"
        errors_str = "; ".join(rec.errors) if rec.errors else ""
        warns_str = "; ".join(rec.warnings) if rec.warnings else ""

        ws_val.cell(row_val, 1, value=rec.user_tag or rec.raw_tag)
        ws_val.cell(row_val, 2, value=rec.acce_item_symbol)
        ws_val.cell(row_val, 3, value=status)
        ws_val.cell(row_val, 4, value=errors_str)
        ws_val.cell(row_val, 5, value=warns_str)
        row_val += 1

    if 'Contents' in wb.sheetnames:
        ws_cont = wb['Contents']
        wb.move_sheet(ws_cont, 0)
        # Очищаем старые ссылки
        for row in ws_cont.iter_rows(min_row=5, max_col=2, values_only=False):
            for cell in row:
                cell.value = None
                cell.hyperlink = None
    else:
        ws_cont = wb.create_sheet("Contents", 0)

    ws_cont['A1'] = f"Project: {project_name}"
    ws_cont['A2'] = "Generated by ACCE Exporter"
    ws_cont['A4'] = "Sheet Name"
    ws_cont['B4'] = "Link"

    row_cont = 5
    for sheet_name in wb.sheetnames:
        if sheet_name == "Contents":
            continue
        ws_cont.cell(row_cont, 1, value=sheet_name)
        cell_link = ws_cont.cell(row_cont, 2)
        cell_link.value = f"Go to {sheet_name}"
        cell_link.hyperlink = f"#'{sheet_name}'!A1"
        cell_link.style = "Hyperlink"
        row_cont += 1

    # Сохранение
    wb.save(output_path)
    logger.info(f"Export saved to {output_path}")

    all_warnings = tag_warnings
    if skipped_count > 0:
        all_warnings.append(f"{skipped_count} records skipped due to validation errors.")

    return ExportResult(
        exported_count=len(valid_records),
        skipped_count=skipped_count,
        by_type=by_type_counts,
        output_path=output_path,
        warnings=all_warnings
    )